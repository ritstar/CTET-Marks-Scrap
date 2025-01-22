import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def convert_json_to_excel(json_file, excel_file):
    # Load JSON data
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Sort data by roll number in ascending order
    data.sort(key=lambda x: int(x["personal_info"].get("Roll No", 0)))

    # Create a set to collect all unique subject names
    all_subjects = set()
    for entry in data:
        for subject in entry["marks_info"]["subjects"]:
            if subject["subject"] == "Mathematics & Science":
                all_subjects.add("Mathematics")
                all_subjects.add("Science")
            else:
                all_subjects.add(subject["subject"])

    # Define headers
    headers = [
        "Roll No",
        "Name",
        "Mother's Name",
        "Father's/Husband's Name",
        "Category",
        "Paper Type",
        "Timestamp"
    ]
    # Add subject-specific headers
    for subject in sorted(all_subjects):
        headers.extend([f"{subject} Marks Obtained", f"{subject} Total Marks"])
    headers.append("Total Percentage")  # Add total percentage column

    # Create a new Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "CTET Results"

    # Write headers to the first row
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Define fill colors for conditional formatting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green

    # Write data to the worksheet
    for row_num, entry in enumerate(data, 2):  # Start from row 2
        row = {
            "Roll No": entry["personal_info"].get("Roll No", ""),
            "Name": entry["personal_info"].get("Name", ""),
            "Mother's Name": entry["personal_info"].get("Mother's Name", ""),
            "Father's/Husband's Name": entry["personal_info"].get("Father's/Husband's Name", ""),
            "Category": entry["personal_info"].get("Category", ""),
            "Paper Type": entry["marks_info"].get("paper_type", ""),
            "Timestamp": entry["timestamp"]
        }

        # Initialize subject marks
        for subject in sorted(all_subjects):
            row[f"{subject} Marks Obtained"] = ""
            row[f"{subject} Total Marks"] = ""

        # Populate subject marks
        total_marks_obtained = 0
        total_marks_total = 0
        for subject in entry["marks_info"]["subjects"]:
            subject_name = subject["subject"]
            marks = subject["marks"]

            if subject_name == "Mathematics & Science":
                # Split into Mathematics and Science
                marks_split = marks.split("(")
                if len(marks_split) > 1:
                    math_science_marks = marks_split[1].strip(")")
                    math_marks, science_marks = math_science_marks.split("Science -")
                    math_marks = math_marks.replace("Mathematics -", "").strip()
                    science_marks = science_marks.strip()

                    row["Mathematics Marks Obtained"] = math_marks
                    row["Mathematics Total Marks"] = "30"  # Assuming total marks for Mathematics is 30
                    row["Science Marks Obtained"] = science_marks
                    row["Science Total Marks"] = "30"  # Assuming total marks for Science is 30

                    total_marks_obtained += int(math_marks) + int(science_marks)
                    total_marks_total += 60  # 30 for Mathematics + 30 for Science
            else:
                if "out of" in marks:
                    marks_obtained, total_marks = marks.split("out of")
                    row[f"{subject_name} Marks Obtained"] = marks_obtained.strip()
                    row[f"{subject_name} Total Marks"] = total_marks.strip()

                    if subject_name.lower() != "total":
                        total_marks_obtained += int(marks_obtained.strip())
                        total_marks_total += int(total_marks.strip())

        # Calculate total percentage
        if total_marks_total > 0:
            total_percentage = (total_marks_obtained / total_marks_total) * 100
            row["Total Percentage"] = f"{total_percentage:.2f}%"
        else:
            row["Total Percentage"] = "N/A"

        # Write row data to the worksheet
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row_num, column=col_num, value=row.get(header, ""))

        # Apply color coding to the Roll No cell based on percentage
        if row["Total Percentage"] != "N/A":
            percentage_value = float(row["Total Percentage"].strip("%"))
            roll_no_cell = ws.cell(row=row_num, column=1)  # Roll No is in the first column
            if percentage_value < 60:
                roll_no_cell.fill = red_fill
            else:
                roll_no_cell.fill = green_fill

    # Save the workbook
    wb.save(excel_file)
    print(f"Excel file saved as {excel_file}")

# Example usage
json_file = "ctet_results.json"
excel_file = "ctet_results.xlsx"
convert_json_to_excel(json_file, excel_file)